import gc
import os
import argparse
from datetime import date, timedelta
from dotenv import load_dotenv
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
EXCEL_EXPORT_LOOKBACK_DAYS = int(os.getenv("EXCEL_EXPORT_LOOKBACK_DAYS", "90"))


def fetch_all(table, filters=None, order=None):
    all_rows = []
    start = 0
    page_size = 1000

    while True:
        query = supabase.table(table).select("*")

        if filters:
            for field, operator, value in filters:
                if operator == "eq":
                    query = query.eq(field, value)
                elif operator == "gte":
                    query = query.gte(field, value)
                elif operator == "gt":
                    query = query.gt(field, value)
                elif operator == "lte":
                    query = query.lte(field, value)

        if order:
            query = query.order(order)

        result = query.range(start, start + page_size - 1).execute()
        rows = result.data or []
        all_rows.extend(rows)

        if len(rows) < page_size:
            break

        start += page_size

    return all_rows


def add_table(ws, title, headers, rows, start_row):
    title_fill = PatternFill("solid", fgColor="EAF2F8")
    header_fill = PatternFill("solid", fgColor="D6EAF8")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)

    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(bold=True, size=14)
    ws.cell(start_row, 1).fill = title_fill

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row + 2, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=start_row + 3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0'

    return start_row + 3 + len(rows)


def style_sheet(ws):
    ws.freeze_panes = "A4"
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.sheet_view.showGridLines = False


def parse_args():
    parser = argparse.ArgumentParser(description="Build management Excel export.")
    parser.add_argument(
        "--full-history",
        action="store_true",
        help="Load full history for heavy source tables instead of the default lookback window.",
    )
    return parser.parse_args()


def build_excel(full_history=False):
    today = date.today()
    date_from_14 = (today - timedelta(days=14)).isoformat()
    lookback_cutoff = (today - timedelta(days=EXCEL_EXPORT_LOOKBACK_DAYS)).isoformat()

    wb = Workbook()
    wb.remove(wb.active)

    wb_daily = fetch_all(
        "daily_marketplace_kpi",
        filters=[("marketplace_code", "eq", "wb"), ("kpi_date", "gte", date_from_14)],
        order="kpi_date"
    )

    ozon_realization = fetch_all(
        "daily_marketplace_kpi",
        filters=[
            ("marketplace_code", "eq", "ozon"),
            ("buyouts_qty", "gt", 0),
            ("buyouts_amount_seller", "gt", 0)
        ],
        order="kpi_date"
    )

    ozon_fbs = fetch_all(
        "marketplace_orders",
        filters=[("marketplace_code", "eq", "ozon"), ("order_date", "gte", date_from_14)],
        order="order_date"
    )

    daily_sku_kpi_filters = None if full_history else [("kpi_date", "gte", lookback_cutoff)]
    daily_sku_kpi = fetch_all("daily_sku_kpi", filters=daily_sku_kpi_filters, order="kpi_date")
    try:
        ozon_organic = fetch_all(
            "ozon_daily_sku_organic",
            filters=[("marketplace_code", "eq", "ozon")],
            order="sale_date",
        )
    except Exception as e:
        print(
            "Не удалось загрузить ozon_daily_sku_organic для Excel. "
            "Продолжаем без листа Ozon Organic SKU. "
            f"Ошибка: {e}"
        )
        ozon_organic = []
    ozon_expenses_filters = [("marketplace_code", "eq", "ozon")]
    if not full_history:
        ozon_expenses_filters.append(("expense_date", "gte", lookback_cutoff))
    ozon_expenses = fetch_all(
        "marketplace_expenses",
        filters=ozon_expenses_filters,
        order="expense_date"
    )

    ws = wb.create_sheet("Summary")
    ws["A1"] = "Управленческий отчет Marketplaces"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A2"] = f"Дата формирования: {today.isoformat()}"

    wb_orders = sum(float(r.get("orders_qty") or 0) for r in wb_daily)
    wb_orders_amount = sum(float(r.get("orders_amount_seller") or 0) for r in wb_daily)
    wb_buyouts = sum(float(r.get("buyouts_qty") or 0) for r in wb_daily)
    wb_buyouts_amount = sum(float(r.get("buyouts_amount_seller") or 0) for r in wb_daily)

    ozon_real_qty = sum(float(r.get("buyouts_qty") or 0) for r in ozon_realization)
    ozon_real_amount = sum(float(r.get("buyouts_amount_seller") or 0) for r in ozon_realization)

    ozon_fbs_qty = sum(float(r.get("orders_qty") or 0) for r in ozon_fbs)
    ozon_fbs_amount = sum(float(r.get("orders_amount_seller") or 0) for r in ozon_fbs)

    summary_rows = [
        ["WB заказы, 14 дней", wb_orders],
        ["WB сумма заказов, 14 дней", wb_orders_amount],
        ["WB выкупы, 14 дней", wb_buyouts],
        ["WB сумма выкупов, 14 дней", wb_buyouts_amount],
        ["Ozon месячная реализация, шт", ozon_real_qty],
        ["Ozon месячная реализация, руб", ozon_real_amount],
        ["Ozon FBS заказы, 14 дней", ozon_fbs_qty],
        ["Ozon FBS сумма заказов, 14 дней", ozon_fbs_amount],
    ]

    add_table(ws, "Ключевые показатели", ["Метрика", "Значение"], summary_rows, 4)
    style_sheet(ws)

    ws_wb = wb.create_sheet("WB Daily")
    wb_rows = []
    for r in wb_daily:
        wb_rows.append([
            r.get("kpi_date"),
            float(r.get("orders_qty") or 0),
            float(r.get("orders_amount_seller") or 0),
            float(r.get("buyouts_qty") or 0),
            float(r.get("buyouts_amount_seller") or 0),
            float(r.get("buyout_rate") or 0),
        ])

    add_table(
        ws_wb,
        "WB: дневная динамика",
        ["Дата", "Заказы", "Сумма заказов", "Выкупы", "Сумма выкупов", "% выкупа"],
        wb_rows,
        1
    )
    style_sheet(ws_wb)

    if len(wb_rows) >= 2:
        chart = LineChart()
        chart.title = "WB: заказы и выкупы"
        chart.y_axis.title = "Шт"
        chart.x_axis.title = "Дата"
        data = Reference(ws_wb, min_col=2, max_col=4, min_row=3, max_row=3 + len(wb_rows))
        cats = Reference(ws_wb, min_col=1, min_row=4, max_row=3 + len(wb_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_wb.add_chart(chart, "H3")

    del wb_daily, wb_rows
    gc.collect()

    ws_ozon = wb.create_sheet("Ozon Realization")
    ozon_rows = []
    for r in ozon_realization:
        ozon_rows.append([
            r.get("kpi_date"),
            float(r.get("buyouts_qty") or 0),
            float(r.get("buyouts_amount_seller") or 0),
            float(r.get("commission_amount") or 0),
        ])

    add_table(
        ws_ozon,
        "Ozon: месячная реализация",
        ["Дата отчета", "Выкупы/шт", "Сумма реализации", "Комиссии"],
        ozon_rows,
        1
    )
    style_sheet(ws_ozon)

    del ozon_realization, ozon_rows
    gc.collect()

    ws_fbs = wb.create_sheet("Ozon FBS")
    grouped_fbs = {}
    for r in ozon_fbs:
        d = r.get("order_date")
        if not d:
            continue
        if d not in grouped_fbs:
            grouped_fbs[d] = [d, 0, 0]
        grouped_fbs[d][1] += float(r.get("orders_qty") or 0)
        grouped_fbs[d][2] += float(r.get("orders_amount_seller") or 0)

    fbs_rows = [grouped_fbs[k] for k in sorted(grouped_fbs.keys())]

    add_table(
        ws_fbs,
        "Ozon FBS: дневные заказы",
        ["Дата", "Заказы FBS", "Сумма заказов FBS"],
        fbs_rows,
        1
    )
    style_sheet(ws_fbs)

    del ozon_fbs, grouped_fbs, fbs_rows
    gc.collect()

    ws_ads = wb.create_sheet("Ozon Ads SKU")
    ads_grouped = {}
    sku_names = {}
    for r in daily_sku_kpi:
        if r.get("marketplace_code") != "ozon":
            continue
        sku = str(r.get("marketplace_sku") or "")
        if not sku:
            continue
        if sku not in sku_names:
            sku_names[sku] = r.get("product_name")

    for r in ozon_expenses:
        expense_type = str(r.get("expense_type") or "")
        if not expense_type.startswith("advertising"):
            continue

        sku = str(r.get("marketplace_sku") or "")
        key = (r.get("expense_date"), sku, r.get("article") or "")

        if key not in ads_grouped:
            ads_grouped[key] = {
                "date": r.get("expense_date"),
                "sku": sku,
                "article": r.get("article") or "",
                "product_name": sku_names.get(sku),
                "advertising_clicks": 0,
                "advertising_order_10": 0,
                "advertising_order_5": 0,
                "advertising_order_other": 0,
                "advertising_order_unknown": 0,
                "advertising_other": 0,
            }

        ads_grouped[key][expense_type] = ads_grouped[key].get(expense_type, 0) + float(r.get("expense_amount") or 0)

    ads_rows = []
    for row in ads_grouped.values():
        total = (
            row["advertising_clicks"]
            + row["advertising_order_10"]
            + row["advertising_order_5"]
            + row["advertising_order_other"]
            + row["advertising_order_unknown"]
            + row["advertising_other"]
        )
        ads_rows.append([
            row["date"],
            row["sku"],
            row["article"],
            row["product_name"],
            row["advertising_clicks"],
            row["advertising_order_10"],
            row["advertising_order_5"],
            row["advertising_order_other"],
            row["advertising_order_unknown"],
            row["advertising_other"],
            total,
        ])

    ads_rows.sort(key=lambda row: (row[0] or "", row[2] or "", row[1] or ""))

    add_table(
        ws_ads,
        "Ozon: рекламные расходы по SKU",
        [
            "Дата",
            "SKU",
            "Артикул",
            "Название",
            "Реклама клики",
            "Оплата за заказ 10%",
            "Оплата за заказ 5%",
            "Оплата за заказ прочие ставки",
            "Оплата за заказ не распознано",
            "Прочая реклама",
            "Всего реклама",
        ],
        ads_rows,
        1
    )
    style_sheet(ws_ads)

    del ozon_expenses, ads_grouped, ads_rows
    gc.collect()

    ws_organic = wb.create_sheet("Ozon Organic SKU")
    organic_rows = []
    for row in ozon_organic:
        organic_rows.append([
            row.get("sale_date"),
            row.get("marketplace_sku"),
            row.get("article"),
            row.get("product_name"),
            float(row.get("total_orders_qty") or 0),
            float(row.get("total_orders_revenue") or 0),
            float(row.get("ad_orders_qty") or 0),
            float(row.get("ad_orders_revenue") or 0),
            float(row.get("organic_orders_qty") or 0),
            float(row.get("organic_orders_revenue") or 0),
            float(row.get("ad_share_orders") or 0),
            float(row.get("ad_share_revenue") or 0),
            row.get("calculation_status"),
            row.get("warning"),
        ])

    add_table(
        ws_organic,
        "Ozon: organic vs ad-attributed sales by SKU",
        [
            "Дата",
            "SKU",
            "Артикул",
            "Название",
            "Total orders qty",
            "Total orders revenue",
            "Ad orders qty",
            "Ad orders revenue",
            "Organic orders qty",
            "Organic orders revenue",
            "Ad share orders",
            "Ad share revenue",
            "Status",
            "Warning",
        ],
        organic_rows,
        1,
    )
    style_sheet(ws_organic)

    del ozon_organic, organic_rows
    gc.collect()

    ws_raw = wb.create_sheet("Raw KPI")
    raw_rows = []
    for r in daily_sku_kpi:
        raw_rows.append([
            r.get("kpi_date"),
            r.get("marketplace_code"),
            r.get("marketplace_sku"),
            r.get("article"),
            r.get("product_name"),
            float(r.get("orders_qty") or 0),
            float(r.get("orders_amount_seller") or 0),
            float(r.get("buyouts_qty") or 0),
            float(r.get("buyouts_amount_seller") or 0),
            float(r.get("buyout_rate") or 0),
            float(r.get("ad_spend") or 0),
            float(r.get("ad_orders_qty") or 0),
            float(r.get("ad_orders_revenue") or 0),
            float(r.get("organic_orders_qty") or 0),
            float(r.get("organic_orders_revenue") or 0),
            float(r.get("ad_share_orders") or 0),
            float(r.get("ad_share_revenue") or 0),
            float(r.get("ad_share_of_orders") or 0),
            float(r.get("roas") or 0),
        ])

    add_table(
        ws_raw,
        "Raw SKU KPI",
        [
            "Дата",
            "MP",
            "SKU",
            "Артикул",
            "Название",
            "Заказы",
            "Сумма заказов",
            "Выкупы",
            "Сумма выкупов",
            "% выкупа",
            "Реклама всего",
            "Ad orders qty",
            "Ad orders revenue",
            "Organic orders qty",
            "Organic orders revenue",
            "Ad share orders",
            "Ad share revenue",
            "ДРР",
            "ROAS",
        ],
        raw_rows,
        1
    )
    style_sheet(ws_raw)

    del daily_sku_kpi, raw_rows
    gc.collect()

    filename = "management_report.xlsx"
    wb.save(filename)
    print(f"✅ Excel-отчет создан: {filename}")


if __name__ == "__main__":
    args = parse_args()
    build_excel(full_history=args.full_history)
