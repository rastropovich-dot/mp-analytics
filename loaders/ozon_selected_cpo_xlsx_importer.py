import argparse
import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime dependency check
    load_workbook = None

load_dotenv()

DATE_KEYS = ("Дата", "date")
SKU_KEYS = ("SKU", "sku", "SKU Ozon")
ARTICLE_KEYS = ("Артикул", "article", "offer_id", "offerId")
NAME_KEYS = ("Название товара", "Название", "product_name", "name")
ORDERS_KEYS = ("Заказы", "Количество", "orders", "qty", "quantity")
REVENUE_KEYS = ("Стоимость продажи, ₽", "Стоимость продажи", "Продажи", "Выручка", "orders_revenue")
RATE_KEYS = ("Ставка, %", "rate_percent", "rate")
SPEND_KEYS = ("Расход, ₽", "Расход", "ad_spend", "spend")
CAMPAIGN_KEYS = ("Кампания", "Название продвижения", "Продвижение", "campaign_name")
CAMPAIGN_ID_KEYS = ("campaign_id", "ID кампании", "Campaign ID")


def parse_args():
    parser = argparse.ArgumentParser(description="Inspect or stage Ozon selected CPO XLSX/CSV files.")
    parser.add_argument("--file", required=True, help="Path to XLSX/CSV exported file")
    parser.add_argument("--date")
    parser.add_argument("--date-from")
    parser.add_argument("--date-to")
    parser.add_argument("--inspect-only", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--write", action="store_true")
    return parser.parse_args()


def num(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text[:10]


def value_by_keys(row, keys):
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def read_csv_rows(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        rows = list(reader)
        return {"sheet_name": "csv", "columns": list(reader.fieldnames or []), "rows": rows}


def read_xlsx_sheets(path):
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to inspect XLSX files. Install it in the runtime used for this importer."
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            sheets.append({"sheet_name": sheet_name, "columns": [], "rows": []})
            continue
        header = [str(v).strip() if v is not None else "" for v in values[0]]
        rows = []
        for raw in values[1:]:
            row = {}
            for idx, cell in enumerate(raw):
                key = header[idx] if idx < len(header) else f"col_{idx}"
                row[key] = cell
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        sheets.append({"sheet_name": sheet_name, "columns": header, "rows": rows})
    return sheets


def detect_sheets(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return [read_csv_rows(path)]
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_sheets(path)
    raise RuntimeError(f"Unsupported file type: {suffix}")


def parse_sheet_rows(sheet, source_file, source_file_hash):
    parsed = []
    for row in sheet["rows"]:
        report_date = normalize_date(value_by_keys(row, DATE_KEYS))
        parsed.append(
            {
                "report_date": report_date,
                "marketplace_code": "ozon",
                "marketplace_sku": str(value_by_keys(row, SKU_KEYS) or "").strip(),
                "article": str(value_by_keys(row, ARTICLE_KEYS) or "").strip(),
                "product_name": str(value_by_keys(row, NAME_KEYS) or "").strip(),
                "campaign_id": str(value_by_keys(row, CAMPAIGN_ID_KEYS) or "").strip(),
                "campaign_name": str(value_by_keys(row, CAMPAIGN_KEYS) or "").strip(),
                "orders_qty": num(value_by_keys(row, ORDERS_KEYS)),
                "orders_revenue": num(value_by_keys(row, REVENUE_KEYS)),
                "rate_percent": num(value_by_keys(row, RATE_KEYS)) or None,
                "ad_spend": num(value_by_keys(row, SPEND_KEYS)),
                "source_file": source_file,
                "source_file_hash": source_file_hash,
                "sheet_name": sheet["sheet_name"],
                "raw_row": row,
            }
        )
    return parsed


def build_summary(sheets, parsed_rows):
    return {
        "sheets": [
            {
                "sheet_name": sheet["sheet_name"],
                "columns": sheet["columns"],
                "row_count": len(sheet["rows"]),
                "sample_rows": sheet["rows"][:5],
            }
            for sheet in sheets
        ],
        "parsed_row_count": len(parsed_rows),
        "total_spend": round(sum(row["ad_spend"] for row in parsed_rows), 2),
        "top_rows": sorted(parsed_rows, key=lambda row: row["ad_spend"], reverse=True)[:10],
    }


def main():
    args = parse_args()
    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        raise RuntimeError(f"File not found: {path}")

    source_file_hash = sha256_file(path)
    sheets = detect_sheets(path)
    parsed_rows = []
    for sheet in sheets:
        parsed_rows.extend(parse_sheet_rows(sheet, str(path), source_file_hash))

    summary = build_summary(sheets, parsed_rows)
    mode = "inspect-only" if args.inspect_only else "dry-run" if args.dry_run else "write" if args.write else "inspect-only"

    print(json.dumps({
        "mode": mode,
        "file": str(path),
        "source_file_hash": source_file_hash,
        "date": args.date,
        "date_from": args.date_from,
        "date_to": args.date_to,
        "summary": summary,
        "writes_marketplace_expenses": False,
        "writes_ozon_daily_sku_ad_attribution": False,
        "writes_only_staging_raw": bool(args.write),
    }, ensure_ascii=False, indent=2, default=str))

    if args.write:
        print("WRITE mode is reserved for future staging-table ingestion. No writes are performed in this step.")


if __name__ == "__main__":
    main()
